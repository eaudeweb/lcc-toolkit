from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from lcc.models import (
    TaxonomyClassification,
    Question,
    Gap,
)


class Command(BaseCommand):

    help = """
        Import Questions and Gaps from an excel file.
        File example in commands/example/original_questions.xlsx
    """

    def process(self, code):
        """
        Updates a code based on the number of previous taxonomy classifications
        sent as parameter. For example, if the questions in a file use taxonomy
        classification codes starting from 1.1, but in our database we already
        have 8 top level classifications, the returned code will be 9.1.
        """
        pt1, pt2 = code.split('.', 1)
        return '.'.join([str(int(pt1) + self.num_prev), pt2])

    def add_arguments(self, parser):
        parser.add_argument('file', type=str)
        # num_prev is the number of previous top-level taxonomy classifications
        # present in the system.
        parser.add_argument("-n", "--num_prev", type=int)
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Only parse the data, but do not insert it."
        )

    def parse_row(self, row):
        # Only return data for rows that have both parent and child
        if row[0].value is None or row[3].value is None:
            return None, None

        parent_code, text = row[0].value.strip().split(' ', 1)
        parent_code = self.process(parent_code)
        
        raw_code, text = row[3].value.strip().split(' ', 1)
        code = self.process(raw_code)
        ret = {
            'text': text,
            'level': 1,
            'parent_answer': 1,
            'classification': code,
            'gap_answer': 0,
            'gap_classifications': [code],
        }
        print(
            "Found classification {} with parent {}".format(
                code, parent_code
            )
        )
        return ret, parent_code

    def create_question(self, data, parent_code, dry_run=False):
        parent_classification = TaxonomyClassification.objects.get(
            code=parent_code
        )
        parent_question = Question.objects.filter(
            classification=parent_classification,
        ).first()
        if not parent_question:
            print("Parent question not found for {}".format(parent_classification))
            return

        classification = TaxonomyClassification.objects.get(
            code=data['classification']
        )
        question = Question.objects.filter(
            classification=classification,
            parent=parent_question
        ).first()
        if question:
            print("Question for {} already created.".format(classification))
            return
        print(
            "Creating question for {} with parent {}".format(
                classification, parent_question
             )
        )
        if not dry_run:
            question = Question.objects.create(
                text=data['text'],
                parent=parent_question,
                parent_answer=data['parent_answer'],
                classification=classification
            )
            return question

    def create_gap(self, data, question, dry_run=False):
        gap_classifications = []
        for code in data['gap_classifications']:
            classification = TaxonomyClassification.objects.get(code=code)
            gap_classifications.append(classification)

        print(
            "Creating gap for question {} with classifications {}".format(
                question, gap_classifications
            )
        )
        if not dry_run:
            gap = Gap.objects.create(on=data['gap_answer'], question=question)
            for classification in gap_classifications:
                gap.classifications.add(classification)

    def handle(self, file, *args, **options):
        wb = load_workbook(file, read_only=True)
        self.num_prev = options.get('num_prev', 0)
        dry_run = options.get('dry_run', False)

        for sheet in wb:
            for row in sheet.iter_rows(min_row=2):
                if not row or not (row[0].value or row[3].value):
                    # No relevant data in this row that cannot be inferred
                    # from other ones, skip it
                    continue
                data, parent_code = self.parse_row(row)
                if data:
                    try:
                        question = self.create_question(data, parent_code, dry_run)
                        if not question:
                            continue
                        self.create_gap(data, question, dry_run)
                    except Exception as e:
                        print(
                            "Failed to create question for {} with error {}".format(
                                data['classification'], str(e)
                            )
                        )
